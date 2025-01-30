from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from csvHandling import Dato

workName = "data.xlsx"

def WB_INIT() -> Workbook:
    try:
        workbook = load_workbook(workName)
    except Exception:
        print("new workbook has made")
        workbook = Workbook()
        workbook.active.title("PC")
        workbook.create_sheet("data")

    workbook.iso_dates = True
    
    return workbook

#header = ["Número PC", "Fecha", "Partida", "Placa", "Procesador", "RAM", "SSD", "Ubicación", "Monitor"]
#worksheet.append(header)


def insertarDato(dato: Dato, workbook: Workbook):
    workbook["PC"].append([dato.numero_pc,
                           dato.fecha,
                           dato.partida,
                           dato.placa,
                           dato.procesador,
                           dato.ram,
                           dato.ssd,
                           dato.ubicacion,
                           dato.monitor])
    workbook.save(workName)

def buscarDatoNumero(numero_pc: int | str, workbook: Workbook) -> int | None:
    for cell in workbook[workName]["A"]:
        if str(cell.value) in str(numero_pc):
            print(f"PC encontrado en fila {cell.row}")
            return cell.row
    print("PC NO encontrado")

def eliminarDato(numero_pc: int | str, workbook: Workbook):
    row = buscarDatoNumero(numero_pc, workbook)
    if row:
        workbook["PC"].delete_rows(row)
