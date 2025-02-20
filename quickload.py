import csvHandling
import openpyxl
import datetime

excel = openpyxl.load_workbook("ListadoPCarmados2024.xlsx")
'''
counter = 0
good = 0

for hoja in excel:
    for row in hoja.iter_rows(values_only=True):
        if None in row:
            print("not good!") 
            counter+=1
        else:
            print("good!")
            good+=1

print(counter)
print(good)
'''
if __name__ == "__main__":
    for hoja in excel:
        firstRow = True
        for row in hoja.iter_rows(values_only=True):
            #If is a valid tuple and isn't the header
            if not None==row[0] and not firstRow:
                data = csvHandling.listToData(row)
                csvHandling.csvInsert(data)
            firstRow = False
                